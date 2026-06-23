#!/usr/bin/env python3
"""Set a grade and/or note for a submitted abstract in the Excel file.

Usage:
    python grade_abstract.py <index> <grade> [<note>]
    python grade_abstract.py <index> --note <note>

    index  1-based abstract number (1 = first data row)
    grade  numeric grade to assign (written to column A)
    note   text note (written to column B, Grade_note)

Example:
    python grade_abstract.py 1 10
    python grade_abstract.py 1 10 "good abstract"
    python grade_abstract.py 1 --note "needs revision"
"""

import sys
from pathlib import Path

import openpyxl

XLSX = Path(__file__).parent / "May 13 Abstracts.xlsx"
HEADER_ROW = 1
COL_GRADE = 1
COL_GRADE_NOTE = 2


def main():
    args = sys.argv[1:]
    if not args:
        print(__doc__.strip())
        sys.exit(1)

    # Parse --note flag
    note = None
    if "--note" in args:
        ni = args.index("--note")
        if ni + 1 >= len(args):
            print("Error: --note requires a value")
            sys.exit(1)
        note = args[ni + 1]
        args = args[:ni] + args[ni + 2:]

    if len(args) < 1 or len(args) > 3:
        print(__doc__.strip())
        sys.exit(1)

    idx = int(args[0])
    grade = float(args[1]) if len(args) >= 2 else None
    if note is None and len(args) == 3:
        note = args[2]

    wb = openpyxl.load_workbook(XLSX)
    ws = wb.active
    max_abstract = ws.max_row - HEADER_ROW

    if idx < 1 or idx > max_abstract:
        print(f"Error: index must be between 1 and {max_abstract}")
        sys.exit(1)

    row = HEADER_ROW + idx  # row 2 for index 1, etc.
    title = ws.cell(row=row, column=10).value or "(no title)"
    authors = ws.cell(row=row, column=12).value or "(no authors)"
    old_grade = ws.cell(row=row, column=COL_GRADE).value
    old_note = ws.cell(row=row, column=COL_GRADE_NOTE).value

    if grade is not None:
        ws.cell(row=row, column=COL_GRADE, value=grade)
    if note is not None:
        ws.cell(row=row, column=COL_GRADE_NOTE, value=note)

    wb.save(XLSX)

    print(f"Abstract {idx}: {title}")
    print(f"  Authors: {authors}")
    if grade is not None:
        print(f"  Grade:   {old_grade} -> {grade}")
    if note is not None:
        print(f"  Note:    {old_note} -> {note}")


if __name__ == "__main__":
    main()
